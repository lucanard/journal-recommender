#!/usr/bin/env python3
"""
Final Merge & Export Script
============================
Takes the fully enriched journal database and produces:
1. Production JSON (clean, deduplicated, scored for completeness)
2. Embedding-ready JSONL (for vector search / semantic matching)
3. Summary statistics and data quality report
4. Updated Excel file with all enrichment data
"""

import json
import os
import sys
import logging
from datetime import datetime

# ─── Config ──────────────────────────────────────────────────────────────────
# Input: use the most enriched version available
POSSIBLE_INPUTS = [
    "enriched_crossref.json",
    "enriched_openalex.json",
    "enriched_doaj.json",
    "journal_database.json",
]

OUTPUT_JSON = "journal_database_final.json"
OUTPUT_JSONL = "journal_embeddings_input.jsonl"
OUTPUT_XLSX = "journal_database_enriched.xlsx"
OUTPUT_REPORT = "enrichment_report.txt"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def compute_completeness(journal):
    """Score 0-100 for how complete a journal record is."""
    fields = {
        "title": 10,
        "publisher": 10,
        "aims_scope": 25,       # Most critical for matching
        "subject_categories": 20,
        "oa_model": 5,
        "apc_display": 5,
        "impact_proxy": 10,
        "indexed_pubmed": 5,
        "in_doaj": 5,
        "homepage": 5,
    }
    score = 0
    for field, weight in fields.items():
        val = journal.get(field)
        if val and val not in [None, "", [], "Unknown", False]:
            score += weight
    return score


def build_embedding_text(journal):
    """
    Construct a single text string for each journal that captures its identity.
    This will be used to generate embeddings for semantic matching.
    """
    parts = []
    
    title = journal.get("title", "")
    if title:
        parts.append(f"Journal: {title}")
    
    scope = journal.get("aims_scope", "")
    if scope:
        parts.append(f"Scope: {scope}")
    
    subjects = journal.get("subject_categories", [])
    if subjects:
        parts.append(f"Subjects: {', '.join(subjects[:15])}")
    
    topics = journal.get("top_topics", [])
    if topics:
        parts.append(f"Topics: {', '.join(topics[:10])}")
    
    keywords = journal.get("keywords", [])
    if keywords:
        parts.append(f"Keywords: {', '.join(keywords[:10])}")
    
    fields = journal.get("fields", [])
    if fields:
        parts.append(f"Fields: {', '.join(fields)}")
    
    publisher = journal.get("publisher", "")
    if publisher:
        parts.append(f"Publisher: {publisher}")
    
    return " | ".join(parts)


def main():
    # Find best input
    input_file = None
    for candidate in POSSIBLE_INPUTS:
        if os.path.exists(candidate):
            input_file = candidate
            break
    
    if not input_file:
        log.error("No input file found!")
        sys.exit(1)
    
    log.info(f"Loading from: {input_file}")
    journals = load_json(input_file)
    log.info(f"Loaded {len(journals)} journals")
    
    # ─── Clean & Deduplicate ───
    seen_issns = set()
    cleaned = []
    duplicates = 0
    
    for j in journals:
        # Skip withdrawn
        if j.get("doaj_withdrawn"):
            continue
        
        # Deduplicate by ISSN
        eissn = j.get("electronic_issn", "")
        pissn = j.get("print_issn", "")
        key = eissn or pissn
        if key and key in seen_issns:
            duplicates += 1
            continue
        if key:
            seen_issns.add(key)
        if eissn:
            seen_issns.add(eissn)
        if pissn:
            seen_issns.add(pissn)
        
        # Compute completeness score
        j["completeness_score"] = compute_completeness(j)
        
        # Build embedding text
        j["embedding_text"] = build_embedding_text(j)
        
        # Normalize fields
        if isinstance(j.get("subject_categories"), list):
            j["subject_categories"] = [s for s in j["subject_categories"] if s]
        
        cleaned.append(j)
    
    # Sort by completeness (best records first)
    cleaned.sort(key=lambda x: x["completeness_score"], reverse=True)
    
    # Re-assign IDs
    for i, j in enumerate(cleaned, 1):
        j["id"] = i
    
    log.info(f"After cleaning: {len(cleaned)} journals ({duplicates} duplicates removed)")
    
    # ─── Save production JSON ───
    save_json(cleaned, OUTPUT_JSON)
    log.info(f"Saved: {OUTPUT_JSON}")
    
    # ─── Save embedding-ready JSONL ───
    with open(OUTPUT_JSONL, "w", encoding="utf-8") as f:
        for j in cleaned:
            if not j.get("embedding_text"):
                continue
            record = {
                "id": j["id"],
                "title": j.get("title", ""),
                "text": j["embedding_text"],
                "issn": j.get("electronic_issn") or j.get("print_issn", ""),
                "indexed_pubmed": j.get("indexed_pubmed", False),
                "in_doaj": j.get("in_doaj", False),
                "oa_model": j.get("oa_model", ""),
                "has_apc": j.get("has_apc"),
                "apc_display": j.get("apc_display", ""),
                "impact_proxy": j.get("impact_proxy", "Unknown"),
                "subject_categories": j.get("subject_categories", []),
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    log.info(f"Saved: {OUTPUT_JSONL}")
    
    # ─── Generate Excel ───
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Enriched Journals"
        
        headers = [
            "ID", "Title", "Publisher", "eISSN", "pISSN",
            "PubMed", "DOAJ", "OA Model", "APC",
            "Impact Proxy", "Subjects", "Fields",
            "Aims & Scope", "Homepage", "Country",
            "H-Index", "Completeness %", "Enrichment Sources"
        ]
        
        hfill = PatternFill("solid", fgColor="1B3A5C")
        hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        dfont = Font(name="Arial", size=9)
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        widths = [6, 40, 25, 12, 12, 8, 8, 12, 15, 10, 35, 30, 50, 30, 8, 8, 10, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        for row_idx, j in enumerate(cleaned, 2):
            values = [
                j["id"],
                j.get("title", ""),
                j.get("publisher", ""),
                j.get("electronic_issn", ""),
                j.get("print_issn", ""),
                "Yes" if j.get("indexed_pubmed") else "No",
                "Yes" if j.get("in_doaj") else "No",
                j.get("oa_model", ""),
                j.get("apc_display", ""),
                j.get("impact_proxy", ""),
                "; ".join(j.get("subject_categories", [])[:5]),
                "; ".join(j.get("fields", [])[:5]),
                (j.get("aims_scope", "") or "")[:200],
                j.get("homepage", ""),
                j.get("country_code", ""),
                j.get("h_index", ""),
                j.get("completeness_score", 0),
                j.get("enrichment_source", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = dfont
            
            # Color by completeness
            score = j.get("completeness_score", 0)
            if score >= 70:
                fill = PatternFill("solid", fgColor="E8F5E9")
            elif score >= 40:
                fill = PatternFill("solid", fgColor="FFF8E1")
            else:
                fill = PatternFill("solid", fgColor="FFEBEE")
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill
        
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cleaned)+1}"
        ws.freeze_panes = "A2"
        
        wb.save(OUTPUT_XLSX)
        log.info(f"Saved: {OUTPUT_XLSX}")
    except ImportError:
        log.warning("openpyxl not available, skipping Excel export")
    
    # ─── Quality Report ───
    total = len(cleaned)
    
    def pct(count):
        return f"{count:,} ({count/total*100:.1f}%)" if total > 0 else "0"
    
    report = f"""
{'='*60}
 JOURNAL DATABASE ENRICHMENT REPORT
 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*60}

OVERVIEW
  Total active journals:  {total:,}
  Duplicates removed:     {duplicates}

INDEXING COVERAGE
  PubMed/MEDLINE:         {pct(sum(1 for j in cleaned if j.get('indexed_pubmed')))}
  DOAJ (Open Access):     {pct(sum(1 for j in cleaned if j.get('in_doaj')))}
  Both:                   {pct(sum(1 for j in cleaned if j.get('indexed_pubmed') and j.get('in_doaj')))}

DATA COMPLETENESS
  Has aims & scope:       {pct(sum(1 for j in cleaned if j.get('aims_scope')))}
  Has subject categories: {pct(sum(1 for j in cleaned if j.get('subject_categories')))}
  Has publisher:          {pct(sum(1 for j in cleaned if j.get('publisher')))}
  Has APC info:           {pct(sum(1 for j in cleaned if j.get('apc_display')))}
  Has impact proxy:       {pct(sum(1 for j in cleaned if j.get('impact_proxy') and j['impact_proxy'] != 'Unknown'))}
  Has homepage URL:       {pct(sum(1 for j in cleaned if j.get('homepage')))}

COMPLETENESS DISTRIBUTION
  High (70-100%):         {pct(sum(1 for j in cleaned if j.get('completeness_score', 0) >= 70))}
  Medium (40-69%):        {pct(sum(1 for j in cleaned if 40 <= j.get('completeness_score', 0) < 70))}
  Low (0-39%):            {pct(sum(1 for j in cleaned if j.get('completeness_score', 0) < 40))}

ENRICHMENT SOURCES
  DOAJ API:               {pct(sum(1 for j in cleaned if 'doaj' in j.get('enrichment_source', '')))}
  OpenAlex:               {pct(sum(1 for j in cleaned if 'openalex' in j.get('enrichment_source', '')))}
  CrossRef:               {pct(sum(1 for j in cleaned if 'crossref' in j.get('enrichment_source', '')))}

IMPACT PROXY DISTRIBUTION
  Q1 (High):              {pct(sum(1 for j in cleaned if j.get('impact_proxy') == 'Q1 (High)'))}
  Q1-Q2:                  {pct(sum(1 for j in cleaned if j.get('impact_proxy') == 'Q1-Q2'))}
  Q2-Q3:                  {pct(sum(1 for j in cleaned if j.get('impact_proxy') == 'Q2-Q3'))}
  Q3-Q4:                  {pct(sum(1 for j in cleaned if j.get('impact_proxy') == 'Q3-Q4'))}
  Q4:                     {pct(sum(1 for j in cleaned if j.get('impact_proxy') == 'Q4'))}
  Unknown:                {pct(sum(1 for j in cleaned if j.get('impact_proxy', 'Unknown') == 'Unknown'))}

FILES PRODUCED
  {OUTPUT_JSON}           — Production database (JSON)
  {OUTPUT_JSONL}          — Embedding input (JSONL, one record per line)
  {OUTPUT_XLSX}           — Excel export (color-coded by completeness)
  {OUTPUT_REPORT}         — This report

NEXT STEPS FOR PRODUCTION
  1. Generate embeddings from {OUTPUT_JSONL} using an embedding model
  2. Store embeddings in a vector database (Pinecone, Weaviate, pgvector, etc.)
  3. Wire up the recommendation API:
     a. Embed user abstract → vector search → top N candidates
     b. Filter by constraints (indexing, OA, APC, article type)
     c. Re-rank with LLM for scope fit explanations
{'='*60}
"""
    
    with open(OUTPUT_REPORT, "w", encoding="utf-8") as f:
        f.write(report)
    
    print(report)


if __name__ == "__main__":
    main()
